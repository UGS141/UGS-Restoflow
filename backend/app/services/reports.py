import io
from datetime import datetime
from typing import List, Dict, Any

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# OpenPyXL imports for Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ReportGenerator:
    """
    Handles PDF and Excel compilations for UGS-Restoflow financial audits.
    """
    @staticmethod
    def generate_sales_pdf(sales_records: List[Dict[str, Any]], filters: Dict[str, Any]) -> io.BytesIO:
        """
        Generates a publication-quality PDF invoice ledger report.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter,
            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        # Custom styles matching clean Apple/Stripe look
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#18181b') # zinc-900
        )
        meta_style = ParagraphStyle(
            'ReportMeta',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#71717a') # zinc-500
        )
        
        story = []
        
        # 1. Header
        story.append(Paragraph("UGS-Restoflow Sales Audit Ledger", title_style))
        story.append(Spacer(1, 6))
        
        # Meta info
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_text = (
            f"Generated At: {now_str} | "
            f"Date Range: {filters.get('start_date', 'All')} to {filters.get('end_date', 'All')} | "
            f"Branch: {filters.get('branch_id', 'All')}"
        )
        story.append(Paragraph(meta_text, meta_style))
        story.append(Spacer(1, 20))
        
        # 2. Main data Table
        table_data = [
            ["Invoice No", "Date", "Waiter", "Payment Method", "Subtotal", "Tax Total", "Grand Total"]
        ]
        
        total_sales = 0.0
        total_tax = 0.0
        
        for record in sales_records:
            total_sales += record.get("grand_total", 0.0)
            total_tax += record.get("tax_total", 0.0)
            
            created_at_raw = record.get("offline_created_at") or record.get("synced_at") or ""
            date_str = created_at_raw[:10] if isinstance(created_at_raw, str) else created_at_raw.strftime("%Y-%m-%d") if created_at_raw else ""
            
            table_data.append([
                record.get("bill_number", "N/A"),
                date_str,
                record.get("cashier_id", "N/A")[:10],
                record.get("payment_method", "N/A").upper(),
                f"INR {record.get('subtotal', 0.0):.2f}",
                f"INR {record.get('tax_total', 0.0):.2f}",
                f"INR {record.get('grand_total', 0.0):.2f}"
            ])
            
        # Summary Row
        table_data.append([
            "TOTALS", "", "", "",
            f"INR {total_sales - total_tax:.2f}",
            f"INR {total_tax:.2f}",
            f"INR {total_sales:.2f}"
        ])
        
        # Table Styling matching clean neutrals
        sales_table = Table(table_data, colWidths=[85, 75, 80, 85, 70, 70, 75])
        sales_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#09090b')), # Header bg (zinc-950)
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#fafafa')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#fcfcfc')),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e4e4e7')), # zinc-200 lines
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fafafa')),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#09090b')),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        story.append(sales_table)
        
        # Build document
        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_sales_excel(sales_records: List[Dict[str, Any]], filters: Dict[str, Any]) -> io.BytesIO:
        """
        Generates a custom-formatted Excel sheet for financial spreadsheets.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Audit"
        
        # Styles
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        header_fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid") # zinc-900
        summary_fill = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid") # zinc-100
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # 1. Report Title Blocks
        ws.append(["UGS-Restoflow - Enterprise Sales Report"])
        ws.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True)
        
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([f"Generated At: {now_str} | Date Range: {filters.get('start_date', 'All')} to {filters.get('end_date', 'All')}"])
        ws.cell(row=2, column=1).font = Font(name="Segoe UI", size=9, italic=True)
        ws.append([]) # empty space
        
        # 2. Table headers
        headers = ["Invoice No", "Date", "Waiter", "Payment Method", "Subtotal (INR)", "Tax Total (INR)", "Grand Total (INR)"]
        ws.append(headers)
        
        # Format Headers
        header_row = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_left
            
        # 3. Add sales items
        start_data_row = 5
        row_idx = start_data_row
        
        for record in sales_records:
            created_at_raw = record.get("offline_created_at") or record.get("synced_at") or ""
            date_str = created_at_raw[:10] if isinstance(created_at_raw, str) else created_at_raw.strftime("%Y-%m-%d") if created_at_raw else ""
            
            ws.append([
                record.get("bill_number", "N/A"),
                date_str,
                record.get("cashier_id", "N/A"),
                record.get("payment_method", "N/A").upper(),
                record.get("subtotal", 0.0),
                record.get("tax_total", 0.0),
                record.get("grand_total", 0.0)
            ])
            row_idx += 1
            
        # 4. Summary Formulas Row
        ws.append([
            "TOTALS", "", "", "",
            f"=SUM(E{start_data_row}:E{row_idx-1})",
            f"=SUM(F{start_data_row}:F{row_idx-1})",
            f"=SUM(G{start_data_row}:G{row_idx-1})"
        ])
        
        summary_row = row_idx
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=summary_row, column=col_idx)
            cell.font = bold_font
            cell.fill = summary_fill
            if col_idx >= 5:
                cell.alignment = align_right
                cell.number_format = '"₹"#,##0.00'
            else:
                cell.alignment = align_left
                
        # Format regular cells
        for r in range(start_data_row, summary_row):
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = regular_font
                if c >= 5:
                    cell.alignment = align_right
                    cell.number_format = '"₹"#,##0.00'
                else:
                    cell.alignment = align_left

        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
